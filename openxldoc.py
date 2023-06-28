import openpyxl

# opening the workbook using the openpyxl module
wb = openpyxl.load_workbook('Data.xlsx')

# identifying the workbook
print(type(wb))

# identifying sheetnames embedded in the workbook - returns a list type
sheetnames_in_wb = wb.sheetnames
print(sheetnames_in_wb)

# isolating a sheet in the current workbook - returns a worksheet object of class "Worksheet"
sheet_in_wb = wb['Sheet1']
print(sheet_in_wb)

# obtaining sheet title - returns a string type
sheet_title = sheet_in_wb.title
print(sheet_title)

