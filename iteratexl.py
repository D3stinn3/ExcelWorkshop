import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb['Sheet1']

class Iterator:
    
    # Add the variable classes to the Iterator class
    def __init__(self):
        self.workbook = wb
        self.sheet = sheet
        
    def loops(self):
        # Initialize loop here and specify the row range using a two step parameter
        # Also specify the seed cell
        for row_ in range(1, 20, 2):
            print(row_, self.sheet.cell(row=row_, column=7).value)
        print("\n")
        
        # Initialize loop here and specify the column range using a two step parameter
        # Also specify the seed cell  
        for column_ in range(1, 8, 2):
            print(column_, self.sheet.cell(row=19, column=column_).value)
        print("\n")
        
        
        # Lets base the loop off limits
        row_lower_limit = self.sheet.min_row
        row_upper_limit = self.sheet.max_row
        column_lower_limit = self.sheet.min_column
        column_upper_limit = self.sheet.max_column
        
        # Iterate using the row limits
        for row_ in range(row_lower_limit, row_upper_limit):
            print(row_, self.sheet.cell(row=row_, column=7).value)
        print("\n")
        
        # Iterate using the column limits  
        for column_ in range(column_lower_limit, column_upper_limit):
            print(column_, self.sheet.cell(row=19, column=column_).value)
        print("\n")
        
    def converter(self):
        # Lets convert cell numbers to letters
        # Translate column 7 to a letter
        print(get_column_letter(7))
        
        # To Access the same via specific cell inputs
        print(self.sheet['G19'].column_letter)
        print(self.sheet.cell(row=19, column=7).column_letter)

        # To verify
        using_column_letter = get_column_letter(3)
        using_column_sheetcell = self.sheet['C3'].column_letter
        using_column_sheetcell2 = self.sheet.cell(row=19, column=3).column_letter
        
        print("\n")
        
        if using_column_letter == using_column_sheetcell and using_column_sheetcell2:
            print("Interconversion is possible between letter and number")
            print("\n")   
        else:
            print("Interconversion between letter and number is impossible")
            print("\n")
    
    def rows_vs_colums(self):
        # Lets see how many rows and columns we have
        print("Rows: ", self.sheet.max_row)
        print("Columns: ", self.sheet.max_column)
        print("\n")
        
        # Closing in on the rows and columns
        # Slicing is one way to do it to get the rectangular area between cell A3 to G5
        Cell_list_A3_to_G2 = list(self.sheet['A3':'G5'])
        print(Cell_list_A3_to_G2)
        print("\n")
        
        # Depends on your worksheet, identify a names column by accessing 
        name_list = list(self.sheet.columns)[6]
        
        # Also depends on your worksheet, identify the rows intended same as the columns
        title_list = list(self.sheet.rows)[1]
        
        # Loop over all the cells to obtain value and coordinates
        for cellobjects in Cell_list_A3_to_G2:
            for cellobject in cellobjects:
                print(cellobject.column_letter, cellobject.coordinate, cellobject.value)
            print("\n")
        
        # Loop over the desired namelist, verifying the cell coordinate in the process    
        for cellobject in name_list:
            print(cellobject.coordinate, cellobject.value)
        print("\n")
        
        # Loop over the desired titlelist, verifyng the cell coordinates in the process
        for cellobj in title_list:
            print(cellobj.coordinate, cellobj.value)
        print("\n")
        

    def list_search(self):
        # Listing out individual cell to search
        specific_cell = list(self.sheet.columns)[1][3]
        print(specific_cell.value)
        print("\n")
        
        # Verifying if cell is contained in a bracket or specific boundaries
        cell_value = specific_cell.value
        location_to_search = self.sheet["A1" : "G7"]
        
        # Not found in area because the cell value is not specified
        if cell_value in location_to_search:
            print("Cell is in the query area")
            print("\n")
        else:
            print("Cell is not in the query area")
            print("\n")
            
        # Verbose search will identify the cell by value    
        for cell_group in location_to_search:
            for cell in cell_group:
                if cell_value == cell.value:
                    print(f"Found it at {cell.coordinate}")
                    
                else:
                    print(f"Not found at {cell.coordinate}") 

        
        
           
if __name__ == "__main__":
    Iterator().loops()
    Iterator().converter()
    Iterator().rows_vs_colums()
    Iterator().list_search()