import openpyxl

# Obtain a single sheet from Workbook
wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb['Sheet1']

# Following up to see if sheet is active or any other active sheets
print(wb.active)

# Get the cell value of A1 from active sheet in current workbook
cell_A1 = sheet['A1']
A1_value = cell_A1.value
print(A1_value)

# Get the cell value of B1 from active sheet in current workbook
cell_B2 = sheet['B2']
B2_value = cell_B2.value
print(B2_value)

# Get the cell value of C3 from active sheet in current workbook
cell_C3 = sheet['C3']
C3_value =cell_C3.value
print(C3_value)

# Get particular rows, columns, and value from the cell

# For A1, B2 and B3
cell_A1_row = cell_A1.row
cell_B2_column = cell_B2.column
cell_C3_value = cell_C3.value


def superposition_A1() -> str:
    # Get the cell value of A1 from active sheet in current workbook
    cell_A1 = sheet['A1']
    
    # Get the cell value, cell row position and cell column position for A1
    cell_A1_value = cell_A1.value
    cell_A1_row = cell_A1.row
    cell_A1_column = cell_A1.column
    
    # Output as a string value
    superposition = "\nA1 cell postion is as follows:\nValue is %s\nRow is %s\nColumn is %s" % (cell_A1_value, cell_A1_row, cell_A1_column)
    
    return superposition

def superposition_B2() -> str:
    # Get the cell value of B2 from active sheet in current workbook
    cell_B2 = sheet['B2']
    
    # Get the cell value, cell row position and cell column position for call B2
    cell_B2_value = cell_B2.value
    cell_B2_row = cell_B2.row
    cell_B2_column = cell_B2.column
    
    superposition = "\nB2 cell position is as follows:\nValue is %s\nRow is %s\nColumn is %s" % (cell_B2_value, cell_B2_row, cell_B2_column)
    return superposition

def reverse_superposition() -> str:
    # Access the cell position from the cell name
    cell_C3 = sheet['C3']
    
    # Accessing cell C3 cell by column and row cordinates
    cell_C3_reverse = sheet.cell(row=3, column=3)
    
    C3_value = cell_C3.value
    C3_value_reverse = cell_C3_reverse.value
    
    # Find out if the value is similar to identify the cell identity    
    if C3_value == C3_value_reverse:
        return "\nValue is similar"
    else:
        return "\nValue is not similar"

def row_loop():
    # Loop through every row
    
    for i in range(1, 10, 2):
        # Use the reverse cell position to iterate through the rows
        # Get the cell value
        return (i, sheet.cell(row=i, column=3).value)
                 

A1_pos = superposition_A1()
B2_pos = superposition_B2()
C3_pos = reverse_superposition()
Cell_row_loop = row_loop()

if __name__ == "__main__":
    print(A1_pos)
    print(B2_pos)
    print(C3_pos)
    print(Cell_row_loop)


