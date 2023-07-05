import openpyxl
import pprint

# Open workbook and access active sheet
print("Subiri Workbook Inafunguka")
wb = openpyxl.load_workbook("Data.xlsx")
sheet = wb['Sheet1']
nhifData = {}

class Practical:

    # Fill nhifData with each patients nhif number na amount alipaswa kulipa
    def __init__(self):
       self.workbook = wb
       self.sheet = sheet
    
    # Wacha tuverify class initialized values kwanza
    def verifra(self):
        print(self.sheet)
        print("\n")
        
    def NhifData(self):
        # Upper and lower row limits
        lower_row_limit = self.sheet.min_row
        upper_row_limit = self.sheet.max_row
        
        # Upper and lower column limits
        lower_row_limit = self.sheet.min_column
        upper_column_limit = self.sheet.max_column
        
        # Tuanze na for loops to check the nhif ref number na amount
        for row in range(lower_row_limit + 1, upper_row_limit + 1):
            nhif_ref_number = sheet['B' + str(row)].value
            amount_paid = sheet['H' + str(row)].value
            print(nhif_ref_number, amount_paid)
        print("\n")
        
            
            
    
    
if __name__ == "__main__":
    Practical().verifra()
    Practical().NhifData()