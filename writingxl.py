import openpyxl

# Ukitaka kuandika workbook mpya
new_wb = openpyxl.Workbook('New.xlsx')
print(f"Workbook created as: {new_wb}")

# Load workbook to write on
wb = openpyxl.load_workbook('Data.xlsx')
sheet = wb['Sheet1']
print(f"Workbook loaded as: {wb}")
print("\n")

# Tutengeneze dunia ya worksheet yetu hivi
class Creation:
    
    def __init__(self) -> None:
        self.newworkbook = new_wb
        self.workbook = wb
        self.sheet = sheet
    
    # Hii ni repesentror tu    
    def __repr__(self) -> str:
        return str(self.workbook)
    
    def create_wb_sheet(self):
        # Wacha tutengeneze sheet yetu
        sheet_creator = self.workbook.create_sheet(title='CodeCreatedSheet')
        print(f"Sheet created as: {sheet_creator}")
        print("\n")
        
        # Verify hapa Kwanza kwa hii instance
        while sheet_creator:
            sheet_name = self.workbook.sheetnames[1]
            print("Sheet exists as: {0}".format(sheet_name))
            print("\n")
            break
        else:
            print("Null")
            
        All_sheets = self.workbook.sheetnames
        
        # Njia ingine ya kuverify
        if 'CodeCreatedSheet' and 'Sheet1' in All_sheets:
            print("Ziko")
            print("\n")
        else:
            print("Haziko")
            print("\n")

            
    
    # Lets verify if the sheet was created in another instance
    def verify(self):
        print(self.workbook.sheetnames)
        print("\n")

    def writecell(self):
        # Wacha tuandike kwenye cell A1 katika workbook yetu katika kila sheet
        self.sheet.cell(row=1, column=1, value="Hello World")
        self.sheet['A1'] = "Hello World"
        
        
        # Wacha tuverify ukweli wa haya maneno
        if self.sheet.cell(row=1, column=1).value == "Hello World":
            self.sheet['A1'] = "hELLO wORLD"
            print("Value changed")
            print("\n")
            print("New value: {0}".format(self.sheet['A1'].value))
            print("\n")
            
      
    def filterwritten(self):
        # Andika sheetnames zote
        my_sheet_list = self.workbook.sheetnames
        
        # Tumia lambda function kudadisi kama sheet one iko
        # Uses boolean to evaluate
        filter_yangu = filter(lambda x: x.count("Sheet1") == 0, my_sheet_list)
        print(list(filter_yangu))

    
               
        

if __name__ == "__main__":
    Creation().create_wb_sheet()
    Creation().verify()
    Creation().writecell()
    Creation().filterwritten()